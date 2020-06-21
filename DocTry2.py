from openpyxl import *
from tkinter import *

# globally declare wb and sheet variable

# opening the existing excel file
wb = load_workbook("F:\\work\\Book1.xlsx")

# create the sheet object
sheet = wb.active



    # Driver code


if __name__ == "__main__":
    # create a GUI window
    root = Tk()

    # set the background colour of GUI window
    root.configure(background='light green')

    # set the title of GUI window
    root.title("Registration Form")

    # set the configuration of GUI window
    root.geometry("500x300")


    def clickregister():
        # create a Form label
        heading = Label(root, text="Registration Form", bg="light green")

        # create a Name label
        name = Label(root, text="Name", bg="light green")

        # create a Contact No. label
        contact_no = Label(root, text="Contact No.", bg="light green")

        # create a Email id label
        email_id = Label(root, text="Email id", bg="light green")

        # create a address label
        password = Label(root, text="Password", bg="light green")

        # grid method is used for placing
        # the widgets at respective positions
        # in table like structure .
        heading.grid(row=0, column=1)
        name.grid(row=1, column=0)
        contact_no.grid(row=2, column=0)
        email_id.grid(row=3, column=0)
        password.grid(row=4, column=0)

        def excel():
            # resize the width of columns in
            # excel spreadsheet
            sheet.column_dimensions['A'].width = 30
            sheet.column_dimensions['B'].width = 20
            sheet.column_dimensions['C'].width = 40
            sheet.column_dimensions['D'].width = 50

            # write given data to an excel spreadsheet
            # at particular location
            sheet.cell(row=1, column=1).value = "Name"
            sheet.cell(row=1, column=2).value = "Contact Number"
            sheet.cell(row=1, column=3).value = "Email id"
            sheet.cell(row=1, column=4).value = "Password"

        # Function to set focus (cursor)
        def focus1(event):
            # set focus on the course_field box
            contact_no_field.focus_set()

        # Function to set focus
        def focus2(event):
            # set focus on the sem_field box
            email_id_field.focus_set()

        # Function to set focus
        def focus3(event):
            # set focus on the form_no_field box
            password_field.focus_set()

        # Function for clearing the
        # contents of text entry boxes
        def clear():
            # clear the content of text entry box
            name_field.delete(0, END)
            contact_no_field.delete(0, END)
            email_id_field.delete(0, END)
            password_field.delete(0, END)

        # Function to take data from GUI
        # window and write to an excel file
        def insert():
            # if user not fill any entry
            # then print "empty input"
            if (name_field.get() == "" and
                    contact_no_field.get() == "" and
                    email_id_field.get() == "" and
                    password_field.get() == ""):

                print("empty input")

            else:

                # assigning the max row and max column
                # value upto which data is written
                # in an excel sheet to the variable
                current_row = sheet.max_row
                current_column = sheet.max_column

                # get method returns current text
                # as string which we write into
                # excel spreadsheet at particular location
                sheet.cell(row=current_row + 1, column=1).value = name_field.get()
                sheet.cell(row=current_row + 1, column=2).value = contact_no_field.get()
                sheet.cell(row=current_row + 1, column=3).value = email_id_field.get()
                sheet.cell(row=current_row + 1, column=4).value = password_field.get()

                # save the file
                wb.save("F:\\work\\Book1.xlsx")

                # set focus on the name_field box
                name_field.focus_set()

                # call the clear() function
                clear()

                Label(root, text="Registered Successfully", bg="light green").grid(row=6, column=1)

        excel()


        # create a text entry box
        # for typing the information
        name_field = Entry(root)
        contact_no_field = Entry(root)
        email_id_field = Entry(root)
        password_field = Entry(root)

        # bind method of widget is used for
        # the binding the function with the events

        # whenever the enter key is pressed
        # then call the focus2 function
        contact_no_field.bind("<Return>", focus1)

        # whenever the enter key is pressed
        # then call the focus3 function
        email_id_field.bind("<Return>", focus2)

        # whenever the enter key is pressed
        # then call the focus4 function
        password_field.bind("<Return>", focus3)

        # grid method is used for placing
        # the widgets at respective positions
        # in table like structure .
        name_field.grid(row=1, column=1, ipadx="100")
        contact_no_field.grid(row=2, column=1, ipadx="100")
        email_id_field.grid(row=3, column=1, ipadx="100")
        password_field.grid(row=4, column=1, ipadx="100")

        # call excel function
        excel()


        # create a Submit Button and place into the root window
        register['state'] = DISABLED
        submit = Button(root, text="Submit", fg="Black",
                        bg="Red", command=insert)
        submit.grid(row=5, column=1)


    def clickloginwin():
        heading1 = Label(root, text="Login", bg="light green")

        # create a Email id label
        email_id1 = Label(root, text="Email id", bg="light green")

        # create a address label
        password1 = Label(root, text="Password", bg="light green")
        heading1.grid(row=0, column=1)
        email_id1.grid(row=1, column=0)
        password1.grid(row=2, column=0)
        email_id1_field = Entry(root)
        password1_field = Entry(root)
        print(email_id1_field)
        email_id1_field.grid(row=1, column=1, ipadx="100")
        password1_field.grid(row=2, column=1, ipadx="100")
        loginwin['state']=DISABLED
        def clicklogin():

            for i in range(2, 1048577):
                if (sheet.cell(row=i, column=3).value == None):
                    break
                else:
                    if (email_id1_field.get() == sheet.cell(row=i, column=3).value and password1_field.get()==sheet.cell(row=i, column=4).value):

                        Label(root,text="Login Successful", fg="Black", bg="white").grid(row=7, column=1)
                        break
                    else:


                        Label(root, text="Login Failed", fg="Black", bg="white").grid(row=7, column=1)

        login = Button(root, text="Login ", fg="Black", bg="Red", command=clicklogin)
        login.grid(row=5, column=1)




    register = Button(root, text="Register", fg="Black",
                    bg="Red", command=clickregister)
    register.grid(row=5, column=3)
    loginwin = Button(root, text="Login Window", fg="Black",
                      bg="Red", command=clickloginwin)
    loginwin.grid(row=5, column=2)

    # start the GUI
    root.mainloop()

