from tkinter import *
from openpyxl import *

# from tkinter.ttk import *
wb = load_workbook("F:\\work\\BookCalender.xlsx")

# create the sheet object
sheet = wb.active



# Creating master Tkinter window
master = Tk()
master.geometry("175x175")
email_id_field = Entry(master)

# Tkinter string variable
# able to store any string value
v = StringVar(master, "")
values={}
for i in range(2,sheet.max_row+1):
    # Dictionary to create multiple buttons
    values[sheet.cell(row=i, column=1).value]=sheet.cell(row=i, column=2).value

# Loop is used to create multiple Radiobuttons
# rather than creating each button separately
for (text, value) in values.items():
    Radiobutton(master, text=text, variable=v,
                value=value, indicator=0,
                background="light blue").grid()


# Infinite loop can be terminated by
# keyboard or mouse interrupt
# or by any predefined function (destroy())
def excel():
    # resize the width of columns in
    # excel spreadsheet
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['B'].width = 50
    #sheet.cell(row=1, column=1).value = "Time Slot"
    sheet.cell(row=1, column=2).value = "Status"
    sheet.cell(row=1, column=2).value = "Email id"

def focus2(event):
    # set focus on the sem_field box
    email_id_field.focus_set()
def insert():
    # if user not fill any entry
    # then print "empty input"
    if (sheet.cell(row=2,column=2).value =="Booked"):
        Label(master,text="Sorry selected slot is not available").grid(row=3, column=1)
    else:
        email_id=Entry(master)
        sheet.cell(row=2, column=3).value=email_id.get()
        sheet.cell(row=2, column=2).value="Booked"
        Label(master,text="Selected slot Booked").grid()
    excel()
    wb.save("F:\\work\\Book1.xlsx")
book = Button(master, text="Book", fg="Black",
                    bg="Red", command=insert)
book.grid(row=5, column=1)

mainloop()